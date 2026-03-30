from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterator

from django.core.exceptions import ValidationError as DjangoValidationError
from django.core.validators import validate_email
from openpyxl import load_workbook


REQUIRED_HEADERS = ("external_id", "user_id", "email", "subject", "message")
MAX_EXTERNAL_ID_LENGTH = 255
MAX_SUBJECT_LENGTH = 255


class ImportValidationError(Exception):
    pass


@dataclass(slots=True)
class ParsedRow:
    row_number: int
    external_id: str
    user_id: int
    email: str
    subject: str
    message: str


@dataclass(slots=True)
class RowError:
    row_number: int
    message: str


def normalize_headers(header_row: tuple) -> list[str]:
    return [str(cell).strip() if cell is not None else "" for cell in header_row]


def validate_headers(headers: list[str]) -> None:
    missing = [header for header in REQUIRED_HEADERS if header not in headers]
    if missing:
        raise ImportValidationError(
            f"Missing required columns: {', '.join(missing)}"
        )


def parse_row(row_number: int, row: tuple, header_map: dict[str, int]) -> ParsedRow:
    def get_value(name: str):
        index = header_map[name]
        return row[index] if index < len(row) else None

    external_id = str(get_value("external_id") or "").strip()
    user_id_raw = get_value("user_id")
    email = str(get_value("email") or "").strip()
    subject = str(get_value("subject") or "").strip()
    message = str(get_value("message") or "").strip()

    if not external_id:
        raise ImportValidationError("external_id is required")
    if len(external_id) > MAX_EXTERNAL_ID_LENGTH:
        raise ImportValidationError(
            f"external_id is too long (max {MAX_EXTERNAL_ID_LENGTH})"
        )

    if user_id_raw in (None, ""):
        raise ImportValidationError("user_id is required")

    try:
        user_id = int(user_id_raw)
    except (TypeError, ValueError) as exc:
        raise ImportValidationError("user_id must be an integer") from exc

    if user_id <= 0:
        raise ImportValidationError("user_id must be a positive integer")

    if not email:
        raise ImportValidationError("email is required")

    try:
        validate_email(email)
    except DjangoValidationError as exc:
        raise ImportValidationError("email is invalid") from exc

    if not subject:
        raise ImportValidationError("subject is required")
    if len(subject) > MAX_SUBJECT_LENGTH:
        raise ImportValidationError(
            f"subject is too long (max {MAX_SUBJECT_LENGTH})"
        )

    if not message:
        raise ImportValidationError("message is required")

    return ParsedRow(
        row_number=row_number,
        external_id=external_id,
        user_id=user_id,
        email=email,
        subject=subject,
        message=message,
    )


def iter_xlsx_rows(file_path: str | Path) -> Iterator[ParsedRow | RowError]:
    workbook = load_workbook(
        filename=file_path,
        read_only=True,
        data_only=True,
    )

    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)

        try:
            header_row = next(rows)
        except StopIteration as exc:
            raise ImportValidationError("XLSX file is empty") from exc

        headers = normalize_headers(header_row)
        validate_headers(headers)

        header_map = {header: index for index, header in enumerate(headers)}

        for row_number, row in enumerate(rows, start=2):
            if row is None:
                continue

            if all(value is None or str(value).strip() == "" for value in row):
                continue

            try:
                yield parse_row(
                    row_number=row_number,
                    row=row,
                    header_map=header_map,
                )
            except ImportValidationError as exc:
                yield RowError(row_number=row_number, message=str(exc))
    finally:
        workbook.close()